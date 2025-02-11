import streamlit as st
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
from streamlit_sortables import sort_items  # Biblioteca para drag-and-drop

# Definir as colunas padrão
DEFAULT_COLUMNS = [
    "Sample", "Pluton", "Group", "Sub_group", "Rock_type", "Observation", "Age", "Tectonic_setting", "UTM_X", "UTM_Y",
    "Age", "Reference", "Colour", "Symbol", "Size", "SiO2", "TiO2", "Al2O3", "FeO", "FeOt", "Fe2O3", "Fe2O3t", "MnO", "MgO", "CaO", "K2O", "Na2O", "P2O5", "Total", 
    "H2Ot", "LOI", "Li", "Be", "B", "Sc", "V", "Cr", "Ni", "Cu", "Zn", "Rb", "Sr", "Y", 
    "Zr", "Nb", "Cs", "Ba", "La", "Ce", "Pr", "Nd", "Sm", "Eu", "Gd", "Tb", "Dy", "Ho", 
    "Er", "Tm", "Yb", "Lu", "Hf", "Ta", "Pb", "Th", "U", "Co", "Mo", "W", "Ga", "Ge", 
    "As", "In", "Sn", "Sb", "Cd"
]

st.title("Reorganizador de Colunas para Dados Geoquímicos")

# Adicionar assinatura no rodapé
st.markdown("---")
st.markdown(
    """
    **Desenvolvido por [Pedro Armond](https://www.researchgate.net/profile/Pedro-Armond)**  
    📧 E-mail: [pedro.armond@aluno.ufop.edu.br](mailto:pedro.armond@aluno.ufop.edu.br)  
    🌐 ResearchGate: [https://www.researchgate.net/profile/Pedro-Armond](https://www.researchgate.net/profile/Pedro-Armond)
    """
)

# Instruções de uso
st.markdown(
    """
    ### Instruções de Uso:
    1. **Carregue um arquivo Excel (.xlsx):** Utilize a funcionalidade de upload para carregar o arquivo contendo seus dados geoquímicos.
    2. **Selecione a aba desejada:** Caso seu arquivo contenha várias abas, você poderá selecionar a aba que deseja reorganizar.
    3. **Visualize as colunas excedentes:** O algoritmo identificará colunas com labels que não correspondem ao padrão e as mostrará na tela. Essas colunas serão adicionadas ao final da planilha processada.
    4. **Organize as colunas existentes:** Utilize a funcionalidade de drag-and-drop para reorganizar as colunas definidas pelo padrão e os rótulos adicionais (se houver):
        ```
        "Sample", "Pluton", "Group", "Sub_group", "Rock_type", "Observation", "Age", "Tectonic_setting", "UTM_X", "UTM_Y",
        "Age", "Reference", "Colour", "Symbol", "Size", "SiO2", "TiO2", "Al2O3", "FeO", "FeOt", 
        "Fe2O3", "Fe2O3t", "MnO", "MgO", "CaO", "K2O", "Na2O", "P2O5", "Total", "H2Ot", "LOI",
        "Li", "Be", "B", "Sc", "V", "Cr", "Ni", "Cu", "Zn", "Rb", "Sr", "Y", "Zr", "Nb", "Cs", 
        "Ba", "La", "Ce", "Pr", "Nd", "Sm", "Eu", "Gd", "Tb", "Dy", "Ho", "Er", "Tm", "Yb", 
        "Lu", "Hf", "Ta", "Pb", "Th", "U", "Co", "Mo", "W", "Ga", "Ge", "As", "In", "Sn", 
        "Sb", "Cd"
        ```
    5. **Colunas ausentes:** Caso sua planilha não contenha algumas dessas colunas, o algoritmo criará essas colunas automaticamente e preencherá os valores com `None` (equivalente a `NaN`), podendo ser excluídas posteriormente, se desejado.
    6. **Baixe o arquivo reorganizado:** O algoritmo salvará um novo arquivo Excel, preservando o original, e o disponibilizará para download.
    """
)

# Inicializa a lista de rótulos adicionais na session_state, se ainda não existir
if "additional_labels" not in st.session_state:
    st.session_state.additional_labels = []

# Opção opcional para adicionar novos rótulos
if st.checkbox("Deseja adicionar novos rótulos?"):
    novo_rotulo = st.text_input("Digite o novo rótulo:")
    if st.button("Adicionar Rótulo"):
        if novo_rotulo and novo_rotulo not in st.session_state.additional_labels:
            st.session_state.additional_labels.append(novo_rotulo)
            st.success(f"Rótulo '{novo_rotulo}' adicionado com sucesso!")
        elif novo_rotulo in st.session_state.additional_labels:
            st.warning("Esse rótulo já foi adicionado.")
        else:
            st.warning("Por favor, insira um rótulo válido.")
    if st.session_state.additional_labels:
        st.write("Rótulos adicionais atuais:")
        st.write(st.session_state.additional_labels)

# Combina os rótulos padrão com os adicionais (caso haja)
colunas_para_reordenar = DEFAULT_COLUMNS + st.session_state.additional_labels

# Upload da planilha
uploaded_file = st.file_uploader("Carregue sua planilha (.xlsx)", type=["xlsx"])

if uploaded_file:
    # Exibir as abas disponíveis no arquivo Excel
    workbook = pd.ExcelFile(uploaded_file)
    sheet_names = workbook.sheet_names
    st.write("A planilha contém as seguintes abas:")
    selected_sheet = st.selectbox("Selecione a aba que contém os dados:", sheet_names)

    # Leitura da aba selecionada
    df = pd.read_excel(uploaded_file, sheet_name=selected_sheet)
    st.write("Pré-visualização dos dados:")
    st.dataframe(df)

    # Identificar colunas excedentes (que não fazem parte do conjunto padrão)
    extra_columns = list(set(df.columns) - set(DEFAULT_COLUMNS))
    if extra_columns:
        st.warning("As seguintes colunas excedentes foram detectadas e serão adicionadas ao final:")
        st.write(extra_columns)

    # Reordenação com Drag-and-Drop
    st.subheader("Reorganize as colunas desejadas:")
    selected_columns = sort_items(
        items=colunas_para_reordenar,
        key="sortable_columns",
        direction="horizontal",  # Direção horizontal para melhor visualização
    )

    st.write("Ordem selecionada:")
    st.write(selected_columns)

    # Reorganizar colunas ao clicar no botão "Salvar"
    if st.button("Salvar Arquivo"):
        # Identificar colunas faltantes e adicioná-las com valor None
        missing_columns = list(set(selected_columns) - set(df.columns))
        for col in missing_columns:
            df[col] = None

        # Reordenar colunas: as que foram arrastadas + as excedentes
        reordered_columns = selected_columns + extra_columns
        df = df[reordered_columns]

        # Salvar arquivo com novo nome
        original_name = uploaded_file.name
        base_name, ext = os.path.splitext(original_name)
        new_file_name = f"{base_name}_modified{ext}"
        df.to_excel(new_file_name, index=False, engine="openpyxl")

        # Aplicar estilo (fonte vermelha) às colunas excedentes
        workbook = load_workbook(new_file_name)
        worksheet = workbook.active

        start_col = len(selected_columns) + 1
        for col_idx in range(start_col, start_col + len(extra_columns)):
            for row_idx in range(2, worksheet.max_row + 1):  # Ignorar o cabeçalho
                worksheet.cell(row=row_idx, column=col_idx).font = Font(color="FF0000")

        workbook.save(new_file_name)

        st.success(f"Arquivo salvo como {new_file_name}.")
        st.write("Baixe o arquivo abaixo:")
        with open(new_file_name, "rb") as file:
            st.download_button(
                label="📥 Baixar Arquivo Modificado",
                data=file,
                file_name=new_file_name
            )